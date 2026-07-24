#!/usr/bin/env python3
"""Convert the placement-test frequency workbook to the app's CSV format."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from nltk.corpus import cmudict
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = ROOT / "datasets" / "placement_1.csv"
DEFAULT_EXAMPLES = ROOT / "scripts" / "data" / "placement_1_examples.csv"

ARPABET_TO_IPA = {
    "AA": "ɑ",
    "AE": "æ",
    "AH": "ʌ",
    "AO": "ɔ",
    "AW": "aʊ",
    "AY": "aɪ",
    "B": "b",
    "CH": "tʃ",
    "D": "d",
    "DH": "ð",
    "EH": "ɛ",
    "ER": "ɜr",
    "EY": "eɪ",
    "F": "f",
    "G": "ɡ",
    "HH": "h",
    "IH": "ɪ",
    "IY": "i",
    "JH": "dʒ",
    "K": "k",
    "L": "l",
    "M": "m",
    "N": "n",
    "NG": "ŋ",
    "OW": "oʊ",
    "OY": "ɔɪ",
    "P": "p",
    "R": "r",
    "S": "s",
    "SH": "ʃ",
    "T": "t",
    "TH": "θ",
    "UH": "ʊ",
    "UW": "u",
    "V": "v",
    "W": "w",
    "Y": "j",
    "Z": "z",
    "ZH": "ʒ",
}

# CMUdict uses American spellings and does not contain these source forms.
PHONETIC_OVERRIDES = {
    "favourite": "/ˈfeɪvərɪt/",
    "behaviour": "/bɪˈheɪvjər/",
    "neighbour": "/ˈneɪbər/",
    "stomachache": "/ˈstʌməkeɪk/",
}

# These entries are genuinely plural in the source meaning, rather than words
# that merely end in "s" (for example means, politics, or mathematics).
PLURAL_WORDS = {"works", "chips", "goods", "jeans"}


def arpabet_to_ipa(phones: list[str]) -> str:
    result: list[str] = []
    for phone in phones:
        match = re.fullmatch(r"([A-Z]+)([012]?)", phone)
        if not match:
            raise ValueError(f"Unsupported ARPABET phone: {phone}")
        symbol, stress = match.groups()
        ipa = ARPABET_TO_IPA[symbol]
        if symbol == "AH" and stress == "0":
            ipa = "ə"
        elif symbol == "ER" and stress == "0":
            ipa = "ər"
        if stress == "1":
            result.append("ˈ")
        elif stress == "2":
            result.append("ˌ")
        result.append(ipa)
    return f"/{''.join(result)}/"


def load_examples(path: Path) -> dict[str, tuple[str, str]]:
    examples: dict[str, tuple[str, str]] = {}
    with path.open(encoding="utf-8", newline="") as file:
        for row_number, row in enumerate(csv.reader(file), start=1):
            if len(row) != 3:
                raise ValueError(
                    f"Example row {row_number} must have 3 columns, got {len(row)}"
                )
            word, example_en, example_cn = (value.strip() for value in row)
            key = word.casefold()
            if not word or key in examples:
                raise ValueError(
                    f"Example row {row_number} has an empty or duplicate word: {word!r}"
                )
            if key not in example_en.casefold():
                raise ValueError(
                    f"Example row {row_number} does not contain {word!r}: {example_en!r}"
                )
            examples[key] = (example_en, example_cn)
    return examples


def convert(source: Path, output: Path, examples_path: Path) -> int:
    pronunciations = cmudict.dict()
    examples = load_examples(examples_path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    seen: set[str] = set()

    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        _, raw_word, raw_frequency, raw_gloss = values[:4]
        if raw_word is None and raw_gloss is None:
            continue
        if raw_word is None or raw_frequency is None or raw_gloss is None:
            raise ValueError(f"Row {row_number} is incomplete: {values!r}")

        word = str(raw_word).strip()
        key = word.casefold()
        if not word or key in seen:
            raise ValueError(f"Row {row_number} has an empty or duplicate word: {word!r}")
        seen.add(key)

        if key in PHONETIC_OVERRIDES:
            phonetic = PHONETIC_OVERRIDES[key]
        elif key in pronunciations:
            phonetic = arpabet_to_ipa(pronunciations[key][0])
        else:
            raise ValueError(f"No pronunciation found for {word!r}")

        gloss = str(raw_gloss).strip().replace(",", "，")
        if key in PLURAL_WORDS and "（复数）" not in gloss:
            gloss += "（复数）"

        if key not in examples:
            raise ValueError(f"No curated example found for {word!r}")
        example_en, example_cn = examples[key]
        rows.append([word, gloss, phonetic, example_en, example_cn, ""])

    unused_examples = set(examples) - seen
    if unused_examples:
        preview = ", ".join(sorted(unused_examples)[:5])
        raise ValueError(f"Curated examples contain words not in the workbook: {preview}")

    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as file:
        csv.writer(file, lineterminator="\n").writerows(rows)
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path, help="source .xlsx file")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--examples", type=Path, default=DEFAULT_EXAMPLES)
    args = parser.parse_args()
    count = convert(
        args.source.expanduser(),
        args.output.expanduser(),
        args.examples.expanduser(),
    )
    print(f"Imported {count} words into {args.output}")


if __name__ == "__main__":
    main()
